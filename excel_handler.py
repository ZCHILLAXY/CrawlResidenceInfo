"""
Excel Handler
Made with ❤️by Z🐻
"""
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelHandler:

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def read_parent_info(self, file_path: str) -> List[Dict[str, str]]:
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在 {file_path}")

        try:
            df = pd.read_excel(file_path, engine=None, dtype=str)

            headers = [str(col).strip() for col in df.columns]
            self.logger.info(f"检测到表头 {headers}")

            if '姓名' not in headers and 'name' not in headers:
                raise ValueError("Excel should have '姓名' or 'name' column")
            if '身份证号' not in headers and 'pid' not in headers:
                raise ValueError("Excel should have '身份证号' or 'pid' column")

            parent_info_list = []

            name_col = '姓名' if '姓名' in headers else 'name'
            pid_col = '身份证号' if '身份证号' in headers else 'pid'

            for idx, row in df.iterrows():
                name = str(row[name_col]) if pd.notna(row[name_col]) else ''
                name = name.strip()

                pid = str(row[pid_col]) if pd.notna(row[pid_col]) else ''
                pid = pid.strip()

                if not name or not pid:
                    self.logger.warning(f"The {idx + 2} data is not complete, Skip.")
                    continue

                if len(pid) not in [15, 18]:
                    self.logger.warning(f"The pid format on {idx + 2} is wrong: {pid}")
                    continue

                parent_info_list.append({
                    'name': name,
                    'pid': pid,
                    'row_number': idx + 2
                })

            self.logger.info(f"Successfully loaded {len(parent_info_list)} infos.")
            return parent_info_list

        except FileNotFoundError:
            raise FileNotFoundError(f"No Existing File: {file_path}")
        except ValueError as e:
            raise ValueError(str(e))
        except Exception as e:
            self.logger.error(f"Read Excel file failed: {str(e)}")
            raise ValueError(f"Invalid file or format error: {str(e)}")

    def write_results(self, output_path: str, parent_info_list: List[Dict[str, str]], query_results: List[Dict[str, Any]]) -> None:
        if len(parent_info_list) != len(query_results):
            raise ValueError("Parent info and search result are not matched.")

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Result"

            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            normal_font = Font(name='微软雅黑', size=10)
            normal_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            headers = ['序号', '姓名', '身份证号', '查询状态', '积分信息', '错误信息']
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row_idx, (parent_info, query_result) in enumerate(zip(parent_info_list, query_results), start=2):
                # 序号
                cell = sheet.cell(row=row_idx, column=1)
                cell.value = row_idx - 1
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # 姓名
                cell = sheet.cell(row=row_idx, column=2)
                cell.value = parent_info['name']
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = border

                # 身份证号
                cell = sheet.cell(row=row_idx, column=3)
                pid = parent_info['pid']
                masked_pid = pid[:6] + '****' + pid[-4:] if len(pid) >= 10 else pid
                cell.value = masked_pid
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = border

                # 查询状态
                cell = sheet.cell(row=row_idx, column=4)
                status = query_result.get('status', 'failed')
                status_text = {
                    'success': '成功',
                    'failed': '失败',
                    'not_found': '未找到记录'
                }.get(status, '未知')
                cell.value = status_text
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                if status == 'success':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif status == 'failed':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

                cell = sheet.cell(row=row_idx, column=5)
                if status == 'success' and query_result.get('data'):
                    data = query_result['data']
                    if 'raw_text' in data:
                        cell.value = data['raw_text']
                    else:
                        cell.value = str(data)
                else:
                    cell.value = ''
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = border

                cell = sheet.cell(row=row_idx, column=6)
                cell.value = query_result.get('error', '')
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = border

            column_widths = {
                'A': 8,
                'B': 12,
                'C': 18,
                'D': 12,
                'E': 60,
                'F': 30
            }

            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width

            sheet.freeze_panes = 'A2'

            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            workbook.close()

            self.logger.info(f"The results have been saved to: {output_path}")

        except Exception as e:
            self.logger.error(f"Write into Excel file failed: {str(e)}")
            raise

    def validate_excel_file(self, file_path: str) -> tuple[bool, Optional[str]]:
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                return False, "文件不存在"

            if file_path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
                return False, "不支持文件格式，请用.xlsx或.xls文件"

            try:
                df = pd.read_excel(file_path, engine=None, nrows=1)

                if df.empty:
                    return False, "文件内容为空"

                headers = [str(col).strip() for col in df.columns]

                if '姓名' not in headers and 'name' not in headers:
                    return False, "缺少'姓名'列"

                if '身份证号' not in headers and 'pid' not in headers:
                    return False, "缺少'身份证号'列"

                return True, None

            except openpyxl.utils.exceptions.InvalidFileException:
                return False, "无效的Excel文件"

        except Exception as e:
            return False, f"验证失败 {str(e)}"

